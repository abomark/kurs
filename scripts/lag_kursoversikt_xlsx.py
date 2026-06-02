"""Genererer KURSOVERSIKT.xlsx fra modul-oversikten.

Engangs-/hjelpeskript: bygger en pen Excel-tabell (rekkefølge, seksjon, tema,
beskrivelse, tilhørende oppgave) i kursets palett (jf. DESIGN_GUIDE §2).
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# Palett (DESIGN_GUIDE §2)
MARINE = "0A2C72"
SAND = "F8E6D5"
HVIT = "FFFFFF"

RADER = [
    (1, "Introduksjon", "Evolusjon", "Bakteppe: utviklingen fram mot agentisk koding/Cortex Code", "-"),
    (2, "Introduksjon", "Cortex Code", "Hva Cortex Code er; enkelt-agent der du sier hva du vil, og agenten velger stegene", "-"),
    (3, "Introduksjon", "Snowsight vs CLI", "To måter å kjøre på: terminal (CLI) vs nettleser (Snowsight); fordeler/ulemper/når", "-"),
    (4, "Introduksjon", "Cortex Code i Snowsight", "Cortex Code i Snowflakes web-UI konkret", "-"),
    (5, "Komme i gang", "Arkitekturoversikt", "Enkelt-agent-arkitektur: en modell styrt av system-prompt med fem lag", "-"),
    (6, "Komme i gang", "Første demo", "Live-demo: workspace, første interaksjon, smakebit AGENTS.md, kostnadsdashbord", "-"),
    (7, "Komme i gang", "Individuell oppgave 1", "Første hands-on, deltakerne prøver selv", "Individuell oppgave"),
    (8, "Komme i gang", "Kostnader", "Gjennomgang i PowerPoint (plenum)", "-"),
    (9, "@-mentions", "@-mentions", "@ som eksplisitt katalog-referanse (autocomplete, chips, hva agenten ser)", "-"),
    (10, "@-mentions", "Individuell oppgave: @-mentions", "Øv på @-referanser", "Individuell oppgave"),
    (11, "Plan Mode", "Plan Mode", "La agenten planlegge før den utfører", "-"),
    (12, "Plan Mode", "Individuell oppgave: Plan Mode", "Øv på Plan Mode", "Individuell oppgave"),
    (13, "AGENTS.md", "AGENTS.md", "Konfigurasjonsfil som styrer agentens kontekst/atferd", "-"),
    (14, "AGENTS.md", "Gruppeoppgave 1", "Interaktiv workshop (Supabase-basert, ordsky + barchart av svar)", "Gruppeoppgave"),
    (15, "AGENTS.md", "Resultater Gruppeoppgave 1", "Samlet resultatvisning for gruppeoppgave 1", "Resultatside"),
    (16, "Modellvalg", "Tilgjengelige modeller", "Hvilke modeller finnes og forskjeller", "-"),
    (17, "Modellvalg", "Individuell oppgave: Modellvalg", "Øv på å velge riktig modell", "Individuell oppgave"),
    (18, "skills.md", "skills.md", "Skills-systemet: gjenbrukbare ferdigheter for agenten", "-"),
    (19, "skills.md", "Demo: Bundled skill", "Mønster: forstå skill-en, anvend på levende objekt, kombiner med Plan Mode", "-"),
    (20, "skills.md", "Individuell oppgave: Bundled skill", "Prøv en bundled skill selv", "Individuell oppgave"),
    (21, "skills.md", "Gruppeoppgave 2", "Gruppeoppgave knyttet til skills", "Gruppeoppgave"),
    (22, "memory.md", "memory.md", "Persistent minne for agenten", "-"),
    (23, "memory.md", "Gruppeoppgave 3", "Gruppeoppgave knyttet til memory", "Gruppeoppgave"),
    (24, "memory.md", "Resultater Gruppeoppgave 3", "Samlet resultatvisning for gruppeoppgave 3", "Resultatside"),
    (25, "Context engineering", "Context engineering", "Hvordan gi agenten riktig kontekst", "-"),
    (26, "Anvendt praksis", "Individuell oppgave 2", "Anvendt øving", "Individuell oppgave"),
    (28, "Dybde", "Demo 2", "Dypere demo", "-"),
    (30, "Dybde", "Autonomous loop", "Den autonome sløyfen i detalj (eks: månedsrapport over kundefrafall - agenten sjekker rolle/RBAC, finner tabeller, skriver+kjører+validerer SQL)", "-"),
    (32, "KURS-data", "Individuell oppgave: Kohortanalyse", "Hands-on på KURS-datasett", "Individuell oppgave"),
    (34, "KURS-data", "Gruppeoppgave: Konkurrent-signaler", "Finn konkurrent-signaler i data", "Gruppeoppgave"),
    (35, "Avslutning", "Avslutning", "Oppsummering, 3-5 take-aways", "-"),
    (36, "Test", "Test: Skills (HTML)", "Intern testside, ikke del av kursløpet", "-"),
]

HEADERE = ["Nr", "Seksjon", "Tema", "Kort beskrivelse", "Tilhørende oppgave"]
BREDDER = [6, 20, 34, 60, 22]

wb = Workbook()
ws = wb.active
ws.title = "Kursoversikt"

kant = Border(*(4 * (Side(style="thin", color="D9D9D9"),)))

# Header-rad
for c, tittel in enumerate(HEADERE, start=1):
    celle = ws.cell(row=1, column=c, value=tittel)
    celle.font = Font(name="Arial", bold=True, color=HVIT, size=11)
    celle.fill = PatternFill("solid", fgColor=MARINE)
    celle.alignment = Alignment(vertical="center", horizontal="left", wrap_text=True)
    celle.border = kant
ws.row_dimensions[1].height = 24

# Datarader (sand-stripet annenhver)
for r, rad in enumerate(RADER, start=2):
    for c, verdi in enumerate(rad, start=1):
        celle = ws.cell(row=r, column=c, value=verdi)
        celle.font = Font(name="Arial", size=10)
        celle.alignment = Alignment(vertical="top", horizontal="left", wrap_text=True)
        celle.border = kant
        if r % 2 == 0:
            celle.fill = PatternFill("solid", fgColor=SAND)

# Kolonnebredder
for i, b in enumerate(BREDDER, start=1):
    ws.column_dimensions[get_column_letter(i)].width = b

ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:E{len(RADER) + 1}"

ut = "KURSOVERSIKT.xlsx"
wb.save(ut)
print(f"Skrev {ut} med {len(RADER)} rader")
